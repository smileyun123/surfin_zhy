#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: zhanghaiyun
@Time: 2025/7/17 14:14  
"""



def read_txt_by_line(file_path):
    """
    读取本地 txt 文件，并以换行符分割每一行内容。

    :param file_path: str, 文件路径
    :return: list, 每个元素为文件中的一行内容
    """
    with open(file_path, 'r') as file:
        lines = file.read().splitlines()
    return lines


lines = read_txt_by_line('../db/fea_off/files/all_odi_output.txt')
print(lines)

# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: zhanghaiyun
@Time: 2026/3/18 16:18  
"""
import json
import os
import argparse
from openpyxl import load_workbook
import csv


def read_excel_with_openpyxl(excel_file, sheet_name=None):
    """
    使用 openpyxl 读取 Excel 文件

    Args:
        excel_file: Excel 文件路径
        sheet_name: 工作表名称，如果为 None 则读取第一个工作表

    Returns:
        list: 包含字典的列表，每个字典代表一行数据，键为列名
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"文件不存在：{excel_file}")

    wb = load_workbook(filename=excel_file, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在，可用的工作表: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    data = []
    headers = None

    for row in ws.iter_rows():
        if headers is None:
            headers = [cell.value for cell in row]
        else:
            row_data = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = cell.value
            data.append(row_data)

    wb.close()
    return data


def read_excel_all_sheets(excel_file):
    """
    读取 Excel 文件中所有工作表的数据

    Args:
        excel_file: Excel 文件路径

    Returns:
        dict: 键为工作表名称，值为该工作表的数据列表
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"文件不存在：{excel_file}")

    wb = load_workbook(filename=excel_file, read_only=True, data_only=True)
    all_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        headers = None

        for row in ws.iter_rows():
            if headers is None:
                headers = [cell.value for cell in row]
            else:
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = cell.value
                data.append(row_data)

        all_data[sheet_name] = data

    wb.close()
    return all_data


def xlsx_to_csv_with_openpyxl(xlsx_file, csv_file=None, sheet_name=None):
    """
    使用 openpyxl 将 xlsx 文件转换为 csv 文件

    Args:
        xlsx_file: xlsx 文件路径
        csv_file: 输出的 csv 文件路径，如果为 None，则默认使用相同文件名
        sheet_name: 要转换的工作表名称，如果为 None 则转换第一个工作表

    Returns:
        str: 转换后的 csv 文件路径
    """
    if not os.path.exists(xlsx_file):
        raise FileNotFoundError(f"文件不存在：{xlsx_file}")

    if csv_file is None:
        base_name = os.path.splitext(xlsx_file)[0]
        csv_file = f"{base_name}.csv"

    data = read_excel_with_openpyxl(xlsx_file, sheet_name)

    if not data:
        print("Excel 文件中没有数据")
        return None

    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    print(f"转换成功：{xlsx_file} -> {csv_file}")
    return csv_file


def batch_convert_xlsx_to_csv_with_openpyxl(directory, output_dir=None):
    """
    批量转换目录下的所有 xlsx 文件为 csv

    Args:
        directory: 包含 xlsx 文件的目录
        output_dir: 输出目录，如果为 None，则保存在原目录

    Returns:
        list: 转换成功的文件列表
    """
    if not os.path.isdir(directory):
        print "error"

    if output_dir is None:
        output_dir = directory

    os.makedirs(output_dir)

    converted_files = []
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') and not filename.startswith('~'):
            xlsx_path = os.path.join(directory, filename)
            csv_path = os.path.join(output_dir, os.path.splitext(filename)[0] + '.csv')

            try:
                xlsx_to_csv_with_openpyxl(xlsx_path, csv_path)
                converted_files.append(csv_path)
            except Exception as e:
                print "ddd"

    return converted_files

# ... existing code ...
